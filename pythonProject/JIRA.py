
import requests
import json
from openpyxl import Workbook
from openpyxl.styles import Font





# Base encode email and api token
cred =  "Bearer "+"OTkxNTIwNDE5Njk0Ok5dRMHFFazdMkDadfkzwLUQ1J5F"
# Set header parameters
headers = {
   "Accept": "application/json",
   "Content-Type": "application/json",
   "Authorization" : cred
}

requested = 0
search_results = dict()
total = 1
jira_items = {
    "Key" : [],
    "Type" : [],
    "Created_Time" : [],
    "Creator" : [],
    "Status" : [],
    "Summary" : [],
    "FixVersions":[],
    "Changelog":[],
    "Assignee": [],
    "Reporter":[]
}


while requested < total:
    post_data = json.dumps({'jql': 'project = ARTBCRC AND issuetype in ("Fault Report", Story) AND fixVersion = PI_22w22 AND "Leading Work Group" in ("ART - BCRC - Domain", "ART - BCRC - BSW Diag and Com", "ART - BCRC - SysSW CI", "ART - BCRC - BSW SW Platform and BL", "ART - BCRC - SysSW System and database", "ART - BCRC - FSW2", "ART - BCRC - FSW", "ART - BCRC - SysSW System Safety and Security", "ART - BCRC - BSW HW Interface")', "startAt": requested, "maxResults": 100})
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Authorization": cred
    }
    response = requests.post('https://jira-vira.volvocars.biz/rest/api/2/search?', headers=headers, data=post_data)
    data = json.loads(response.text)
    total = data['total']
    search_results[requested] = data
    requested += 100
    print("connection")
for base in search_results:
    #for key in search_results[base]['issues']:
       #print(key['key'])
    for item in search_results[base]['issues']:
        jira_items["Key"].append(item["key"])
        jira_items["Type"].append(item["fields"]["issuetype"]["name"])
        jira_items["Created_Time"].append(item["fields"]["created"])
        jira_items["Creator"].append(item["fields"]["creator"]["displayName"])
        jira_items["Status"].append(item["fields"]["status"]["name"])
        jira_items["Summary"].append(item["fields"]["summary"])


wb = Workbook()
ws = wb.active

key_row = 2
issuetype_row = 2
created_time_row = 2
creator_row = 2
status_row = 2
summary_row = 2
changelog_row=2


start_column = 1
header_column = 1
# adding header
headerList = ['Issue Key', 'Issue Type', 'Created Time', 'Creator', 'Status', 'Summary']
for x in headerList:
    ws.cell(row=1, column=header_column).value = x
    header_column += 1
font = Font(color="FF0000")
ws.freeze_panes = "A2"
ws["A1"].font = font
ws["B1"].font = font
ws["C1"].font = font
ws["D1"].font = font
ws["E1"].font = font
ws["F1"].font = font
ws.print_title_rows = '1:1'
print(f'There is total {total} issues')

for key in jira_items["Key"]:
    ws.cell(row=key_row, column=start_column).value = key
    key_row += 1
for issuetype in jira_items["Type"]:
    ws.cell(row=issuetype_row, column=start_column + 1).value = issuetype
    issuetype_row += 1
print("debug3")
for created in jira_items["Created_Time"]:
    ws.cell(row=created_time_row, column=start_column + 2).value = created
    created_time_row += 1
for creator in jira_items["Creator"]:
    ws.cell(row=creator_row, column=start_column + 3).value = creator
    creator_row += 1
for status in jira_items["Status"]:
    ws.cell(row=status_row, column=start_column + 4).value = status
    status_row += 1
for summary in jira_items["Summary"]:
    ws.cell(row=summary_row, column=start_column + 5).value = summary
    summary_row += 1
for changelog in jira_items["Changelog"]:
    ws.cell (row=changelog_row, column=start_column+6).value=changelog
    changelog_row += 1
print("Items are added to Excel File!")

wb.save("jira-report.xlsx")
print("Excel File is saved!")








"""
import requests
from openpyxl import Workbook
from openpyxl.styles import Font
import json

# Base encode email and api token
cred =  "Bearer "+"OTkxNTIwNDE5Njk0Ok5dRMHFFazdMkDadfkzwLUQ1J5F"
# Set header parameters
headers = {
   "Accept": "application/json",
   "Content-Type": "application/json",
   "Authorization" : cred
}

# Enter your project key here
projectKey = "ARTBCRC"
data_list=[]
issue_type_list = []
created_time_list = []
creator_list = []
status_list = []
summary_list = []

jira_type = {
    "Key" : [],
    "Type" : []
}

def get_all_issues():
    issues =[]
    i = 0
    runs = 0
    maxResult = 50
    index = 0
    count = 0
    # Update your site url
    url = "https://jira-vira.volvocars.biz/rest/api/2/search?jql=project%20%3D%20ARTBCRC%20AND%20issuetype%20in%20(%22Fault%20Report%22%2C%20Story)%20AND%20fixVersion%20%3D%20PI_22w22%20AND%20%22Leading%20Work%20Group%22%20in%20(%22ART%20-%20BCRC%20-%20Domain%22%2C%20%22ART%20-%20BCRC%20-%20BSW%20Diag%20and%20Com%22%2C%20%22ART%20-%20BCRC%20-%20SysSW%20CI%22%2C%20%22ART%20-%20BCRC%20-%20BSW%20SW%20Platform%20and%20BL%22%2C%20%22ART%20-%20BCRC%20-%20SysSW%20System%20and%20database%22%2C%20%22ART%20-%20BCRC%20-%20FSW2%22%2C%20%22ART%20-%20BCRC%20-%20FSW%22%2C%20%22ART%20-%20BCRC%20-%20SysSW%20System%20Safety%20and%20Security%22%2C%20%22ART%20-%20BCRC%20-%20BSW%20HW%20Interface%22)&expand=changelog"
    #url_final = url + "&startindex=" + str(index) + "&maxResults=" + str(maxResult)
    url_final = "https://jira-vira.volvocars.biz/rest/api/2/search?jql=project%20%3D%20ARTBCRC%20AND%20issuetype%20in%20(%22Fault%20Report%22%2C%20Story)%20AND%20fixVersion%20%3D%20PI_22w22%20AND%20%22Leading%20Work%20Group%22%20in%20(%22ART%20-%20BCRC%20-%20Domain%22%2C%20%22ART%20-%20BCRC%20-%20BSW%20Diag%20and%20Com%22%2C%20%22ART%20-%20BCRC%20-%20SysSW%20CI%22%2C%20%22ART%20-%20BCRC%20-%20BSW%20SW%20Platform%20and%20BL%22%2C%20%22ART%20-%20BCRC%20-%20SysSW%20System%20and%20database%22%2C%20%22ART%20-%20BCRC%20-%20FSW2%22%2C%20%22ART%20-%20BCRC%20-%20FSW%22%2C%20%22ART%20-%20BCRC%20-%20SysSW%20System%20Safety%20and%20Security%22%2C%20%22ART%20-%20BCRC%20-%20BSW%20HW%20Interface%22)&startIndex=100"
    #print(url_final)
    print('debug 1')
    response = requests.request("Get", url_final, headers=headers)
    json_data = json.loads(response.content)
    size = json_data["total"]
    runs = int(size/maxResult)
    print('debug 2')
    while i <= runs:
        print('debug')


        #url_temp = url + "&startIndex=" + str(index) + "&maxResults=" + str(maxResult)
        url_temp = "https://jira-vira.volvocars.biz/rest/api/2/search?jql=project%20%3D%20ARTBCRC%20AND%20issuetype%20in%20(%22Fault%20Report%22%2C%20Story)%20AND%20fixVersion%20%3D%20PI_22w22%20AND%20%22Leading%20Work%20Group%22%20in%20(%22ART%20-%20BCRC%20-%20Domain%22%2C%20%22ART%20-%20BCRC%20-%20BSW%20Diag%20and%20Com%22%2C%20%22ART%20-%20BCRC%20-%20SysSW%20CI%22%2C%20%22ART%20-%20BCRC%20-%20BSW%20SW%20Platform%20and%20BL%22%2C%20%22ART%20-%20BCRC%20-%20SysSW%20System%20and%20database%22%2C%20%22ART%20-%20BCRC%20-%20FSW2%22%2C%20%22ART%20-%20BCRC%20-%20FSW%22%2C%20%22ART%20-%20BCRC%20-%20SysSW%20System%20Safety%20and%20Security%22%2C%20%22ART%20-%20BCRC%20-%20BSW%20HW%20Interface%22)&startIndex=100"
        response = requests.request("Get", url_temp, headers=headers)
        json_data = json.loads(response.text)

        print(index)

        for item in json_data["issues"]:
            jira_type["Key"].append(item["key"])
            jira_type["Type"].append(item["fields"]["issuetype"]["name"])

        i += 1
        index = index + maxResult

    return json_data

created_time_list.append(item["fields"]["created"])creator_list.append(item["fields"]["creator"]["displayName"])
status_list.append(item["fields"]["status"]["name"])
summary_list.append(item["fields"]["summary"])


get_all_issues()


print(response.status_code)
#print(response.text)

# Decode Json string to Python
json_data = json.loads(response.text)
print(json_data)

key_list = []
issue_type_list = []
created_time_list=[]
creator_list=[]
status_list=[]
summary_list=[]
change_log_list=[]

for item in json_data["issues"]:
    print(item["id"] + "\t" + item["key"] + "\t" +
        item["fields"]["issuetype"]["name"] + "\t" +
        item["fields"]["created"]+ "\t" +
        item["fields"]["creator"]["displayName"] + "\t" +
        item["fields"]["status"]["name"] + "\t" +
        item["fields"]["summary"] + "\t"
        )

for item in json_data["issues"]:
   key_list.append(item["key"])
   issue_type_list.append(item["fields"]["issuetype"]["name"])
   created_time_list.append(item["fields"]["created"])
   creator_list.append(item["fields"]["creator"]["displayName"])
   status_list.append(item["fields"]["status"]["name"])
   summary_list.append(item["fields"]["summary"])
   change_log_list.append(item["changelog"]["histories"]["id"])



wb = Workbook()
ws = wb.active
key_row=2
issuetype_row=2
created_time_row=2
creator_row=2
status_row=2
summary_row=2
#changelog_row=2


start_column =1
header_column =1
# adding header
headerList = ['Issue Key', 'Issue Type', 'Created Time','Creator','Status','Summary']
for x in headerList:
    ws.cell (row=1,column=header_column).value = x
    header_column +=1
font = Font(color="FF0000")
ws.freeze_panes = "A2"
ws["A1"].font = font
ws["B1"].font = font
ws["C1"].font = font
ws["D1"].font = font
ws["E1"].font = font
ws["F1"].font = font
ws.print_title_rows = '1:1'


for key in data_list:
       ws.cell(row=key_row, column =start_column).value =key
       key_row += 1
for issuetype in issue_type_list:
       ws.cell (row=issuetype_row, column=start_column+1).value=issuetype
       issuetype_row += 1

for created in created_time_list:
       ws.cell (row=created_time_row, column=start_column+2).value=created
       created_time_row += 1
for creator in creator_list:
       ws.cell (row=creator_row, column=start_column+3).value=creator
       creator_row += 1
for status in status_list:
       ws.cell (row=status_row, column=start_column+4).value=status
       status_row += 1
for summary in summary_list:
       ws.cell (row=summary_row, column=start_column+5).value=summary
       summary_row += 1

#for changelog in change_log_list:
       #ws.cell (row=changelog_row, column=start_column+6).value=changelog
       #changelog_row += 1
       
wb.save("jira-report.xlsx")

"""


